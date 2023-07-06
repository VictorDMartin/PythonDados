import requests
from bs4 import BeautifulSoup
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Pegando acesso ao HTML
headers = {'user-agent': 'Mozilla/5.0'}
url = requests.get('https://www.agostinholeiloes.com.br', headers=headers)
url_html = url.text
html = BeautifulSoup(url_html, 'html.parser')

# Buscando dados dos lotes
lista_empresas = html.find_all('h6', {'class': re.compile('card-title m-0')})
lista_datas = html.find_all('div', {'class': re.compile('card-text mb-auto')})
lista_documento = html.find_all('div', {'class': re.compile('card-body d-flex flex-column')})
lista_imagem = html.find_all('img', {'class': re.compile('my-auto')})

# Criando o dataframe
dados = []
for i, datas in enumerate(lista_datas):
    row = {}
    if i < len(lista_empresas):
        row['Nome da Empresa'] = lista_empresas[i].text

    paragrafos = [dados.text.strip() for dados in datas.find_all('p')]
    row['Dados do Lote'] = '\n'.join(paragrafos)

    if i < len(lista_documento):
        documento = lista_documento[i]
        link = documento.find('a', href=True)
        if link:
            row['Link'] = link['href']

    if i < len(lista_imagem):
        imagem = lista_imagem[i]
        row['Link da Imagem'] = imagem['src']

    dados.append(row)

df = pd.DataFrame(dados)

# Criando a planilha Excel
wb = Workbook()
ws = wb.active

# Definindo estilos de formatação
header_font = Font(bold=True)
cell_alignment = Alignment(wrap_text=True)

# Escrevendo o cabeçalho
for col_num, column_name in enumerate(df.columns, 1):
    cell = ws.cell(row=1, column=col_num, value=column_name)
    cell.font = header_font

# Escrevendo os dados
for row_num, row_data in enumerate(df.values, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num, value=cell_value)
        cell.alignment = cell_alignment

# Ajustando a largura das colunas
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except TypeError:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[column].width = adjusted_width

# Salvando a planilha
wb.save('dados_formatados.xlsx')
